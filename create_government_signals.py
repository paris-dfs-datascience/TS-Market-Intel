#!/usr/bin/env python3
"""
Generate government_signals.xlsx from government_results.json
Thomas Scientific // Government Market Intelligence
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE = "/Users/amanpreet/Desktop/thomas-scientific-market-intelligence/80s-accounts-market-intelligence/backend/accounts-vertical"
INPUT  = os.path.join(BASE, "government_results.json")
OUTPUT = os.path.join(BASE, "government_signals.xlsx")

# ─── Colour palette ───────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "EBF3FB"
LIGHT_GRAY  = "D9D9D9"
WHITE       = "FFFFFF"
GRAY_TITLE  = "808080"

SIGNAL_COLORS = {
    "grant":     "DDEEFF",
    "capital":   "DDFFDD",
    "contract":  "FFFACC",
    "expansion": "D5F5E3",
    "project":   "FFE5CC",
    "tender":    "FFF3CC",
}

SIGNAL_TYPES = ["grant", "capital", "contract", "expansion", "project", "tender"]

# ─── Helper: make a fill ──────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, italic=False, size=10, color="000000", name="Arial"):
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── Load data ────────────────────────────────────────────────────────────────
with open(INPUT, "r", encoding="utf-8") as f:
    data = json.load(f)

# Build summary counts
account_rows = []
for record in data:
    account = record["account"]
    category = record["category"]
    signals  = record.get("signals", {})
    counts   = {st: len(signals.get(st, [])) for st in SIGNAL_TYPES}
    total    = sum(counts.values())
    account_rows.append({
        "account":  account,
        "category": category,
        "counts":   counts,
        "total":    total,
        "signals":  signals,
        "timestamp": record.get("timestamp", ""),
    })

# Sort by total signals descending
account_rows.sort(key=lambda x: x["total"], reverse=True)

# ─── Workbook ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"

# Column widths
ws1.column_dimensions["A"].width = 35
for col_letter in ["B","C","D","E","F","G","H"]:
    ws1.column_dimensions[col_letter].width = 10

# Row 1 – Title (merged A1:H1)
ws1.merge_cells("A1:H1")
title_cell = ws1["A1"]
title_cell.value = "Thomas Scientific // Government Market Intelligence"
title_cell.font  = font(bold=True, size=16, color=WHITE)
title_cell.fill  = fill(DARK_BLUE)
title_cell.alignment = align(h="center", v="center")
ws1.row_dimensions[1].height = 30

# Row 2 – Subtitle (merged A2:H2)
ws1.merge_cells("A2:H2")
sub_cell = ws1["A2"]
sub_cell.value = "Last 7 Days  |  April 7, 2026"
sub_cell.font  = font(italic=True, size=11, color=GRAY_TITLE)
sub_cell.fill  = fill("F2F2F2")
sub_cell.alignment = align(h="center", v="center")
ws1.row_dimensions[2].height = 20

# Row 3 – Blank
ws1.row_dimensions[3].height = 6

# Row 4 – Table header
headers = ["Account", "Total Signals", "Grant", "Capital", "Contract", "Expansion", "Project", "Tender"]
for col_idx, hdr in enumerate(headers, start=1):
    cell = ws1.cell(row=4, column=col_idx, value=hdr)
    cell.font      = font(bold=True, size=10, color=WHITE)
    cell.fill      = fill(DARK_BLUE)
    cell.alignment = align(h="center", v="center")
    cell.border    = thin_border()
ws1.row_dimensions[4].height = 20

# Data rows (start at row 5)
data_start_row = 5
top3 = set()
for i, row_data in enumerate(account_rows):
    if i < 3:
        top3.add(row_data["account"])

for i, row_data in enumerate(account_rows):
    row_num = data_start_row + i
    is_alt  = (i % 2 == 1)
    bg_hex  = LIGHT_BLUE if is_alt else WHITE
    is_top3 = row_data["account"] in top3

    # Column A — Account
    cell_a = ws1.cell(row=row_num, column=1, value=row_data["account"])
    cell_a.font      = font(bold=is_top3, size=10)
    cell_a.fill      = fill(bg_hex)
    cell_a.alignment = align(h="left", v="center")
    cell_a.border    = thin_border()

    # Column B — Total Signals (SUM formula)
    c_start = get_column_letter(3)  # C
    c_end   = get_column_letter(8)  # H
    cell_b = ws1.cell(row=row_num, column=2)
    cell_b.value     = f"=SUM({c_start}{row_num}:{c_end}{row_num})"
    cell_b.font      = font(bold=is_top3, size=10)
    cell_b.fill      = fill(bg_hex)
    cell_b.alignment = align(h="center", v="center")
    cell_b.border    = thin_border()

    # Columns C–H — individual signal type counts
    for col_offset, st in enumerate(SIGNAL_TYPES):
        col_num = 3 + col_offset
        cell = ws1.cell(row=row_num, column=col_num, value=row_data["counts"][st])
        cell.font      = font(bold=is_top3, size=10)
        cell.fill      = fill(bg_hex)
        cell.alignment = align(h="center", v="center")
        cell.border    = thin_border()

    ws1.row_dimensions[row_num].height = 18

# Totals row
total_row = data_start_row + len(account_rows)
ws1.cell(row=total_row, column=1, value="TOTAL").font = font(bold=True, size=10)
ws1.cell(row=total_row, column=1).fill      = fill(LIGHT_GRAY)
ws1.cell(row=total_row, column=1).alignment = align(h="left", v="center")
ws1.cell(row=total_row, column=1).border    = thin_border()

for col_num in range(2, 9):
    col_letter = get_column_letter(col_num)
    cell = ws1.cell(row=total_row, column=col_num)
    cell.value     = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})"
    cell.font      = font(bold=True, size=10)
    cell.fill      = fill(LIGHT_GRAY)
    cell.alignment = align(h="center", v="center")
    cell.border    = thin_border()

ws1.row_dimensions[total_row].height = 20

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — All Signals
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Signals")

# Column widths
col_widths = {
    1: 25,   # Account
    2: 15,   # Category
    3: 12,   # Signal Type
    4: 60,   # Summary
    5: 50,   # Why It Matters
    6: 40,   # Source URL
    7: 20,   # Detail 1
    8: 20,   # Detail 2
    9: 20,   # Detail 3
}
for col_num, width in col_widths.items():
    ws2.column_dimensions[get_column_letter(col_num)].width = width

# Header row
s2_headers = ["Account","Category","Signal Type","Summary","Why It Matters","Source URL","Detail 1","Detail 2","Detail 3"]
for col_idx, hdr in enumerate(s2_headers, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=hdr)
    cell.font      = font(bold=True, size=10, color=WHITE)
    cell.fill      = fill(DARK_BLUE)
    cell.alignment = align(h="center", v="center")
    cell.border    = thin_border()
ws2.row_dimensions[1].height = 20

# Freeze top row
ws2.freeze_panes = "A2"

def get_details(signal_type, sig):
    """Return up to 3 (label: value) strings for the Detail columns."""
    if signal_type == "grant":
        return [
            f"Amount: {sig.get('amount','N/A')}",
            f"Agency: {sig.get('agency','N/A')}",
            f"Recipient: {sig.get('recipient','N/A')}",
        ]
    elif signal_type == "capital":
        return [
            f"Value: {sig.get('value', sig.get('investment_size','N/A'))}",
            f"Location: {sig.get('location','N/A')}",
            f"Round: {sig.get('round_type', sig.get('timeline','N/A'))}",
        ]
    elif signal_type == "contract":
        est = sig.get('estimated_value', sig.get('value','N/A'))
        return [
            f"Value: {est}",
            f"Counterparty: {sig.get('counterparty', sig.get('contract_name','N/A'))}",
            f"Deadline: {sig.get('deadline_or_expiration','N/A')}",
        ]
    elif signal_type == "expansion":
        return [
            f"Location: {sig.get('location','N/A')}",
            f"Investment: {sig.get('investment_value', sig.get('investment_size','N/A'))}",
            f"Type: {sig.get('type_of_expansion', sig.get('facility_type','N/A'))}",
        ]
    elif signal_type == "project":
        return [
            f"Project: {sig.get('project_name','N/A')}",
            f"Timeline: {sig.get('timeline','N/A')}",
            f"Budget: {sig.get('budget', sig.get('scope','N/A'))[:80] if sig.get('budget', sig.get('scope')) else 'N/A'}",
        ]
    elif signal_type == "tender":
        name = sig.get('tender_name', sig.get('tender_title','N/A'))
        return [
            f"Tender: {name}",
            f"Deadline: {sig.get('deadline','N/A')}",
            f"Est. Value: {sig.get('estimated_value','N/A')}",
        ]
    return ["", "", ""]

s2_row = 2
for row_data in account_rows:
    for sig_type in SIGNAL_TYPES:
        for sig in row_data["signals"].get(sig_type, []):
            details = get_details(sig_type, sig)
            bg = SIGNAL_COLORS.get(sig_type, WHITE)

            row_vals = [
                row_data["account"],
                row_data["category"],
                sig_type,
                sig.get("summary",""),
                sig.get("why_it_matters",""),
                sig.get("source_url",""),
                details[0],
                details[1],
                details[2],
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws2.cell(row=s2_row, column=col_idx, value=val)
                cell.fill   = fill(bg)
                cell.border = thin_border()
                if col_idx in (4, 5):
                    cell.alignment = align(h="left", v="top", wrap=True)
                    cell.font = font(size=9)
                elif col_idx == 6:
                    cell.alignment = align(h="left", v="top", wrap=True)
                    cell.font = font(size=8)
                elif col_idx == 3:
                    cell.alignment = align(h="center", v="center")
                    cell.font = font(bold=True, size=9)
                else:
                    cell.alignment = align(h="left", v="top", wrap=True)
                    cell.font = font(size=9)

            ws2.row_dimensions[s2_row].height = 60
            s2_row += 1

# ══════════════════════════════════════════════════════════════════════════════
#  SHEETS 3+ — One per account
# ══════════════════════════════════════════════════════════════════════════════
for row_data in account_rows:
    if row_data["total"] == 0:
        continue

    sheet_name = row_data["account"][:31]
    # Sanitise sheet name
    for ch in r'\/?*[]':
        sheet_name = sheet_name.replace(ch, " ")

    ws = wb.create_sheet(sheet_name)

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25

    current_row = 1

    # Account title
    ws.merge_cells(f"A{current_row}:F{current_row}")
    title = ws.cell(row=current_row, column=1, value=row_data["account"])
    title.font = font(bold=True, size=16, color=WHITE)
    title.fill = fill(DARK_BLUE)
    title.alignment = align(h="left", v="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Subtitle
    ws.merge_cells(f"A{current_row}:F{current_row}")
    sub = ws.cell(
        row=current_row, column=1,
        value=f"{row_data['category']}  |  {row_data['total']} signal(s)  |  {row_data['timestamp'][:10]}"
    )
    sub.font = font(italic=True, size=10, color=GRAY_TITLE)
    sub.fill = fill("F2F2F2")
    sub.alignment = align(h="left", v="center")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for sig_type in SIGNAL_TYPES:
        signals_list = row_data["signals"].get(sig_type, [])
        if not signals_list:
            continue

        # Blank spacer
        ws.row_dimensions[current_row].height = 8
        current_row += 1

        # Section header
        bg_color = SIGNAL_COLORS.get(sig_type, LIGHT_GRAY)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        sec = ws.cell(row=current_row, column=1, value=sig_type.upper())
        sec.font = font(bold=True, size=11, color="333333")
        sec.fill = fill(bg_color)
        sec.alignment = align(h="left", v="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Column headers for this section
        col_hdrs = ["Summary", "Why It Matters", "Source URL", "Detail 1", "Detail 2", "Detail 3"]
        for col_idx, hdr in enumerate(col_hdrs, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=hdr)
            c.font = font(bold=True, size=9, color=WHITE)
            c.fill = fill(DARK_BLUE)
            c.alignment = align(h="center", v="center")
            c.border = thin_border()
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Signal rows
        for i, sig in enumerate(signals_list):
            is_alt = (i % 2 == 1)
            bg = "F8FEFF" if is_alt else WHITE
            details = get_details(sig_type, sig)

            vals = [
                sig.get("summary", ""),
                sig.get("why_it_matters", ""),
                sig.get("source_url", ""),
                details[0],
                details[1],
                details[2],
            ]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.fill   = fill(bg)
                c.border = thin_border()
                if col_idx in (1, 2):
                    c.alignment = align(h="left", v="top", wrap=True)
                    c.font = font(size=9)
                elif col_idx == 3:
                    c.alignment = align(h="left", v="top", wrap=True)
                    c.font = font(size=8)
                else:
                    c.alignment = align(h="left", v="top", wrap=True)
                    c.font = font(size=9)
            ws.row_dimensions[current_row].height = 70
            current_row += 1

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✓ Saved: {OUTPUT}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
print(f"  Accounts: {len(account_rows)}, with signals: {sum(1 for r in account_rows if r['total'] > 0)}")
total_signals = sum(r['total'] for r in account_rows)
print(f"  Total signals across all accounts: {total_signals}")
