import re
import json
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from collections import defaultdict

with open('jhu_grants_all.json') as f:
    raw = json.load(f)

ILLEGAL = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]')

def safe(val, default=''):
    if val is None:
        return default
    if isinstance(val, list):
        return ', '.join(str(v) for v in val)
    if isinstance(val, dict):
        return json.dumps(val)
    if isinstance(val, str):
        return ILLEGAL.sub('', val)
    return val

def flatten(g):
    org = g.get('organization') or {}
    pi_list = g.get('principal_investigators') or []
    pi = pi_list[0] if pi_list else {}
    ic = g.get('agency_ic_admin') or {}
    ss = g.get('full_study_section') or {}
    org_type = g.get('organization_type') or {}
    return {
        'Application ID':       safe(g.get('appl_id')),
        'Project Number':       safe(g.get('project_num')),
        'Core Project Num':     safe(g.get('core_project_num')),
        'Fiscal Year':          safe(g.get('fiscal_year')),
        'Activity Code':        safe(g.get('activity_code')),
        'Award Type':           safe(g.get('award_type')),
        'Funding Mechanism':    safe(g.get('funding_mechanism')),
        'Is Active':            safe(g.get('is_active')),
        'Is New':               safe(g.get('is_new')),
        'Project Title':        safe(g.get('project_title')),
        'Abstract':             safe(g.get('abstract_text')),
        'Public Health Relevance': safe(g.get('phr_text')),
        'Award Amount ($)':     safe(g.get('award_amount'), 0),
        'Direct Cost ($)':      safe(g.get('direct_cost_amt'), 0),
        'Indirect Cost ($)':    safe(g.get('indirect_cost_amt'), 0),
        'PI Full Name':         safe(pi.get('full_name')),
        'PI First Name':        safe(pi.get('first_name')),
        'PI Last Name':         safe(pi.get('last_name')),
        'PI Title':             safe(pi.get('title')),
        'Contact PI':           safe(g.get('contact_pi_name')),
        'Organization':         safe(org.get('org_name')),
        'City':                 safe(org.get('org_city')),
        'State':                safe(org.get('org_state')),
        'Country':              safe(org.get('org_country')),
        'Zip Code':             safe(org.get('org_zipcode')),
        'Dept Type':            safe(org.get('dept_type')),
        'Org Type':             safe(org_type.get('name')),
        'Congressional Dist':   safe(g.get('cong_dist')),
        'Agency Code':          safe(g.get('agency_code')),
        'IC Code':              safe(ic.get('code')),
        'IC Name':              safe(ic.get('name')),
        'IC Abbreviation':      safe(ic.get('abbreviation')),
        'Opportunity Number':   safe(g.get('opportunity_number')),
        'Study Section':        safe(ss.get('name')),
        'Project Start Date':   safe(g.get('project_start_date'), '')[:10],
        'Project End Date':     safe(g.get('project_end_date'), '')[:10],
        'Budget Start':         safe(g.get('budget_start'), '')[:10],
        'Budget End':           safe(g.get('budget_end'), '')[:10],
        'Award Notice Date':    safe(g.get('award_notice_date'), '')[:10],
        'Date Added':           safe(g.get('date_added'), '')[:10],
        'ARRA Funded':          safe(g.get('arra_funded')),
        'COVID Response':       safe(g.get('covid_response')),
        'CFDA Code':            safe(g.get('cfda_code')),
        'Detail URL':           safe(g.get('project_detail_url')),
        'Keywords':             safe(g.get('pref_terms')),
    }

records = [flatten(g) for g in raw]
headers = list(records[0].keys())

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE  = 'FF1F3864'
MID_BLUE   = 'FF2E5090'
LIGHT_BLUE = 'FFD6E4F7'
WHITE      = 'FFFFFFFF'
GREY_ROW   = 'FFF5F7FA'
GOLD       = 'FFFFC000'

hdr_font    = Font(name='Arial', bold=True, color=WHITE, size=10)
hdr_fill    = PatternFill('solid', start_color=DARK_BLUE)
hdr_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
subhdr_fill = PatternFill('solid', start_color=MID_BLUE)
subhdr_font = Font(name='Arial', bold=True, color=WHITE, size=10)
cell_font   = Font(name='Arial', size=9)
alt_fill    = PatternFill('solid', start_color=GREY_ROW)
thin        = Side(style='thin', color='FFB0BEC5')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary Dashboard
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Summary Dashboard'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 18

# Title banner
ws1.merge_cells('A1:F1')
ws1['A1'] = 'Johns Hopkins University — NIH Grants Summary (FY2024–FY2026)'
ws1['A1'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
ws1['A1'].fill = PatternFill('solid', start_color=DARK_BLUE)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:F2')
ws1['A2'] = f'Source: NIH RePORTER API  |  Total Records: {len(records):,}  |  Generated via api.reporter.nih.gov'
ws1['A2'].font = Font(name='Arial', italic=True, size=9, color='FF555555')
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 16

# ── By Fiscal Year ──
r = 4
ws1.merge_cells(f'A{r}:F{r}')
ws1[f'A{r}'] = 'GRANTS BY FISCAL YEAR'
ws1[f'A{r}'].font = subhdr_font
ws1[f'A{r}'].fill = subhdr_fill
ws1[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[r].height = 22
r += 1

for col, hdr in enumerate(['Fiscal Year', 'Grant Count', 'Total Award ($)', 'Avg Award ($)', 'Max Award ($)', 'Min Award ($)'], 1):
    c = ws1.cell(r, col, hdr)
    c.font = Font(name='Arial', bold=True, size=9, color='FF1F3864')
    c.fill = PatternFill('solid', start_color=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='center')
    c.border = border
ws1.row_dimensions[r].height = 18
data_start_row = r + 1

by_year = defaultdict(list)
for rec in records:
    by_year[rec['Fiscal Year']].append(rec['Award Amount ($)'])

for i, year in enumerate(sorted(by_year)):
    r += 1
    amounts = [a for a in by_year[year] if isinstance(a, (int, float)) and a > 0]
    row_fill = PatternFill('solid', start_color=WHITE) if i % 2 == 0 else PatternFill('solid', start_color=GREY_ROW)
    vals = [f'FY{year}', len(by_year[year]),
            sum(amounts), sum(amounts)/len(amounts) if amounts else 0,
            max(amounts) if amounts else 0, min(amounts) if amounts else 0]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(r, col, val)
        c.font = cell_font
        c.fill = row_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if col >= 3:
            c.number_format = '$#,##0'
ws1.row_dimensions[r].height = 16

# Totals row
r += 1
all_amounts = [rec['Award Amount ($)'] for rec in records if isinstance(rec['Award Amount ($)'], (int, float)) and rec['Award Amount ($)'] > 0]
totals = ['TOTAL', f'=SUM(B{data_start_row}:B{r-1})', f'=SUM(C{data_start_row}:C{r-1})', '', '', '']
for col, val in enumerate(totals, 1):
    c = ws1.cell(r, col, val)
    c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_BLUE)
    c.border = border
    c.alignment = Alignment(horizontal='center')
    if col >= 3 and val:
        c.number_format = '$#,##0'
ws1.row_dimensions[r].height = 18

# ── By Activity Code ──
r += 2
ws1.merge_cells(f'A{r}:F{r}')
ws1[f'A{r}'] = 'TOP ACTIVITY CODES'
ws1[f'A{r}'].font = subhdr_font
ws1[f'A{r}'].fill = subhdr_fill
ws1[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[r].height = 22
r += 1

for col, hdr in enumerate(['Activity Code', 'Grant Count', 'Total Award ($)', 'Avg Award ($)', '', ''], 1):
    c = ws1.cell(r, col, hdr)
    c.font = Font(name='Arial', bold=True, size=9, color='FF1F3864')
    c.fill = PatternFill('solid', start_color=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='center')
    c.border = border
ws1.row_dimensions[r].height = 18

by_code = defaultdict(list)
for rec in records:
    by_code[rec['Activity Code']].append(rec['Award Amount ($)'])
top_codes = sorted(by_code.items(), key=lambda x: sum(a for a in x[1] if isinstance(a,(int,float))), reverse=True)[:15]

for i, (code, amounts) in enumerate(top_codes):
    r += 1
    amounts_clean = [a for a in amounts if isinstance(a, (int, float)) and a > 0]
    row_fill = PatternFill('solid', start_color=WHITE) if i % 2 == 0 else PatternFill('solid', start_color=GREY_ROW)
    vals = [code, len(amounts), sum(amounts_clean), sum(amounts_clean)/len(amounts_clean) if amounts_clean else 0, '', '']
    for col, val in enumerate(vals, 1):
        c = ws1.cell(r, col, val)
        c.font = cell_font
        c.fill = row_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if col >= 3 and val != '':
            c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 16

# ── By NIH Institute ──
r += 2
ws1.merge_cells(f'A{r}:F{r}')
ws1[f'A{r}'] = 'TOP NIH INSTITUTES'
ws1[f'A{r}'].font = subhdr_font
ws1[f'A{r}'].fill = subhdr_fill
ws1[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[r].height = 22
r += 1

for col, hdr in enumerate(['Institute', 'Abbreviation', 'Grant Count', 'Total Award ($)', 'Avg Award ($)', ''], 1):
    c = ws1.cell(r, col, hdr)
    c.font = Font(name='Arial', bold=True, size=9, color='FF1F3864')
    c.fill = PatternFill('solid', start_color=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='center')
    c.border = border
ws1.row_dimensions[r].height = 18

by_ic = defaultdict(lambda: {'abbr': '', 'amounts': []})
for rec in records:
    k = rec['IC Name']
    by_ic[k]['abbr'] = rec['IC Abbreviation']
    by_ic[k]['amounts'].append(rec['Award Amount ($)'])
top_ics = sorted(by_ic.items(), key=lambda x: sum(a for a in x[1]['amounts'] if isinstance(a,(int,float))), reverse=True)[:15]

for i, (ic, info) in enumerate(top_ics):
    r += 1
    amounts_clean = [a for a in info['amounts'] if isinstance(a, (int, float)) and a > 0]
    row_fill = PatternFill('solid', start_color=WHITE) if i % 2 == 0 else PatternFill('solid', start_color=GREY_ROW)
    vals = [ic, info['abbr'], len(info['amounts']), sum(amounts_clean), sum(amounts_clean)/len(amounts_clean) if amounts_clean else 0, '']
    for col, val in enumerate(vals, 1):
        c = ws1.cell(r, col, val)
        c.font = cell_font
        c.fill = row_fill
        c.border = border
        c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        if col >= 4 and val != '':
            c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 16

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — All Grants Data
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('All Grants')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A2'

# Header row
for col, hdr in enumerate(headers, 1):
    c = ws2.cell(1, col, hdr)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = border
ws2.row_dimensions[1].height = 30

# Data rows
for row_i, rec in enumerate(records, 2):
    row_fill = PatternFill('solid', start_color=WHITE) if row_i % 2 == 0 else PatternFill('solid', start_color=GREY_ROW)
    for col, key in enumerate(headers, 1):
        val = rec[key]
        c = ws2.cell(row_i, col, val)
        c.font = cell_font
        c.fill = row_fill
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=(key in ('Project Title', 'Abstract', 'Public Health Relevance', 'Keywords')))
        if key in ('Award Amount ($)', 'Direct Cost ($)', 'Indirect Cost ($)'):
            c.number_format = '$#,##0'

# Column widths
col_widths = {
    'Application ID': 14, 'Project Number': 20, 'Core Project Num': 16,
    'Fiscal Year': 10, 'Activity Code': 12, 'Award Type': 10,
    'Funding Mechanism': 20, 'Is Active': 9, 'Is New': 8,
    'Project Title': 55, 'Abstract': 60, 'Public Health Relevance': 50,
    'Award Amount ($)': 16, 'Direct Cost ($)': 16, 'Indirect Cost ($)': 16,
    'PI Full Name': 22, 'PI First Name': 15, 'PI Last Name': 15, 'PI Title': 22,
    'Contact PI': 22, 'Organization': 32, 'City': 14, 'State': 8,
    'Country': 16, 'Zip Code': 12, 'Dept Type': 22, 'Org Type': 24,
    'Congressional Dist': 16, 'Agency Code': 12, 'IC Code': 8,
    'IC Name': 38, 'IC Abbreviation': 14, 'Opportunity Number': 18,
    'Study Section': 45, 'Project Start Date': 16, 'Project End Date': 16,
    'Budget Start': 14, 'Budget End': 14, 'Award Notice Date': 16,
    'Date Added': 14, 'ARRA Funded': 10, 'COVID Response': 14,
    'CFDA Code': 12, 'Detail URL': 45, 'Keywords': 50,
}
for col, key in enumerate(headers, 1):
    ws2.column_dimensions[get_column_letter(col)].width = col_widths.get(key, 14)

# ════════════════════════════════════════════════════════════════════
# SHEET 3 — Top 50 by Award Amount
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Top 50 by Award')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'A2'

top50_headers = ['Rank', 'Project Number', 'Fiscal Year', 'Activity Code', 'Award Amount ($)',
                 'Direct Cost ($)', 'Indirect Cost ($)', 'PI Full Name', 'PI Title',
                 'IC Abbreviation', 'Dept Type', 'Project Title', 'Project Start Date',
                 'Project End Date', 'Detail URL']

for col, hdr in enumerate(top50_headers, 1):
    c = ws3.cell(1, col, hdr)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = border
ws3.row_dimensions[1].height = 30

top50 = sorted(records, key=lambda x: x.get('Award Amount ($)') or 0, reverse=True)[:50]
for row_i, rec in enumerate(top50, 2):
    row_fill = PatternFill('solid', start_color=WHITE) if row_i % 2 == 0 else PatternFill('solid', start_color=GREY_ROW)
    vals = [row_i - 1] + [rec.get(h, '') for h in top50_headers[1:]]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row_i, col, val)
        c.font = cell_font
        c.fill = row_fill
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=(top50_headers[col-1] == 'Project Title'))
        if top50_headers[col-1] in ('Award Amount ($)', 'Direct Cost ($)', 'Indirect Cost ($)'):
            c.number_format = '$#,##0'

top50_widths = [6, 20, 10, 12, 16, 16, 16, 22, 24, 14, 22, 55, 16, 16, 45]
for col, w in enumerate(top50_widths, 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════
out = 'jhu_grants_2024_2026.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Sheets: Summary Dashboard, All Grants ({len(records)} rows), Top 50 by Award')
