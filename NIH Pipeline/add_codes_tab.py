import json, warnings
warnings.filterwarnings('ignore')
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Activity code definitions ───────────────────────────────────────────────
CODES = [
    # (Code, Category, Series, Full Name, Description, Typical Duration, Typical Award Size)
    ("D43",  "Training",              "D-series",  "International Research Training Grant",
     "Builds research capacity at institutions in low- and middle-income countries by funding training programs in biomedical, behavioral, and clinical sciences.",
     "Up to 5 years", "$200K–$500K/yr"),

    ("DP1",  "Director's Awards",     "DP-series", "NIH Director's Pioneer Award",
     "Funds exceptionally creative scientists proposing transformative, high-risk/high-reward research. No preliminary data required. Highly competitive.",
     "5 years", "Up to $700K/yr direct"),

    ("DP2",  "Director's Awards",     "DP-series", "NIH Director's New Innovator Award",
     "Supports early-stage investigators with bold, innovative ideas. Applicants must not have had a prior R01 or equivalent. Skip-the-queue mechanism.",
     "5 years", "Up to $300K/yr direct"),

    ("DP5",  "Director's Awards",     "DP-series", "NIH Director's Early Independence Award",
     "Allows exceptional junior scientists to skip traditional postdoc training and transition directly to independent research positions.",
     "5 years", "Up to $250K/yr direct"),

    ("F30",  "Fellowships",           "F-series",  "Ruth L. Kirschstein Predoctoral MD/PhD Fellowship",
     "Supports dual-degree (MD/PhD or similar) students during the PhD research phase of their training. Covers stipend, tuition, and fees.",
     "Up to 6 years", "~$25K–$57K/yr stipend + tuition"),

    ("F31",  "Fellowships",           "F-series",  "Ruth L. Kirschstein Predoctoral Individual NRSA Fellowship",
     "Supports PhD students conducting dissertation research. Covers stipend, tuition, institutional allowance. One of the most common NIH fellowships.",
     "Up to 5 years", "~$25K–$57K/yr stipend + tuition"),

    ("F32",  "Fellowships",           "F-series",  "Ruth L. Kirschstein Postdoctoral Individual NRSA Fellowship",
     "Supports postdoctoral researchers in gaining advanced training. Covers stipend, tuition if applicable, and institutional allowance.",
     "Up to 3 years", "~$56K–$74K/yr stipend"),

    ("F99",  "Fellowships",           "F-series",  "Predoctoral to Postdoctoral Transition Award (Phase I of F99/K00)",
     "Phase I: supports late-stage PhD students. Upon degree completion, transitions into K00 postdoctoral award. Designed for pathway to independence.",
     "Up to 2 years (F99) + 4 years (K00)", "~$40K–$55K/yr"),

    ("G13",  "Other",                 "G-series",  "Health Sciences Library Improvement Grant",
     "Supports improvements to health sciences libraries serving health professionals and researchers, including digitization and resource access.",
     "1–3 years", "Varies"),

    ("K00",  "Career Development",    "K-series",  "Postdoctoral Transition Award (Phase II of F99/K00)",
     "Phase II of the F99/K00 pathway. Provides salary support and research funds for early postdoctoral career development after completing PhD.",
     "Up to 4 years", "~$90K salary + research costs"),

    ("K01",  "Career Development",    "K-series",  "Mentored Research Scientist Career Development Award",
     "Provides salary support and mentoring for PhD-level researchers who need protected time to develop independent research careers in biomedical fields.",
     "3–5 years", "~$75K–$100K salary + $25K–$50K research"),

    ("K02",  "Career Development",    "K-series",  "Independent Scientist Award",
     "Supports established investigators with a history of NIH funding who need protected research time and wish to pursue new directions.",
     "5 years", "~$100K–$150K salary support"),

    ("K08",  "Career Development",    "K-series",  "Mentored Clinical Scientist Research Career Development Award",
     "For clinically trained MDs/DOs who want to develop into independent laboratory-based researchers. Provides protected research time and mentorship.",
     "3–5 years", "~$100K–$150K salary + research costs"),

    ("K12",  "Career Development",    "K-series",  "Institutional Clinical and Translational Science Award (Career)",
     "Institution-level award that funds multiple trainees. The institution recruits and selects junior faculty scholars to receive career development support.",
     "3 years per scholar", "Varies per scholar"),

    ("K18",  "Career Development",    "K-series",  "Career Enhancement Award",
     "Short-term training support for mid-career investigators who wish to acquire new research skills or shift their research focus.",
     "1–2 years", "Modest research costs"),

    ("K22",  "Career Development",    "K-series",  "Career Transition Award",
     "Supports NCI intramural scientists transitioning to extramural (academic) positions. Provides bridge funding for the first years of independence.",
     "Up to 5 years", "~$250K/yr total"),

    ("K23",  "Career Development",    "K-series",  "Mentored Patient-Oriented Research Career Development Award",
     "For clinician-scientists conducting patient-oriented research (clinical trials, outcomes, epidemiology). Provides protected research time and mentoring.",
     "3–5 years", "~$100K–$150K salary + research costs"),

    ("K24",  "Career Development",    "K-series",  "Midcareer Investigator Award in Patient-Oriented Research",
     "Supports mid-career established investigators so they can devote time to mentoring clinical research trainees while continuing their own research.",
     "5 years", "~$100K–$150K salary support"),

    ("K25",  "Career Development",    "K-series",  "Mentored Quantitative Research Career Development Award",
     "For quantitative scientists (statisticians, engineers, computer scientists, physicists) who want to apply their skills to biomedical/clinical research.",
     "3–5 years", "~$75K–$100K salary + research costs"),

    ("K99",  "Career Development",    "K-series",  "Pathway to Independence Award — Phase I (Mentored)",
     "Phase I of K99/R00. Supports postdoctoral researchers in a mentored setting as they prepare to transition to independent faculty positions. Must transition within 2 years.",
     "Up to 2 years (K99) + 3 years (R00)", "~$90K salary + $21K research"),

    ("N01",  "Contracts",             "N-series",  "Research and Development Contract",
     "NIH directly procures specific goods or services (e.g., large-scale genotyping, animal models, clinical trial coordination). More prescriptive than grants — NIH specifies deliverables.",
     "Varies", "Varies widely (can be $M+)"),

    ("N02",  "Contracts",             "N-series",  "NIH Equipment/Services Contract",
     "NIH contract for procurement of equipment, supplies, or specific technical services needed to support research programs.",
     "Short-term", "Varies"),

    ("OT2",  "Other Transactions",    "OT-series", "Other Transaction Agreement — Phase II",
     "Flexible non-grant, non-contract funding mechanism (authorized under 21st Century Cures Act). Used for innovative projects that don't fit traditional grant structures, e.g. large consortia, data platforms.",
     "Varies", "Varies, can be multi-million"),

    ("P01",  "Center Grants",         "P-series",  "Research Program Project Grant",
     "Funds multiple thematically related research projects under one umbrella PI, with shared cores. Each project has its own PI. High prestige, complex to administer.",
     "5 years", "$1M–$5M+/yr"),

    ("P2C",  "Center Grants",         "P-series",  "Research Center Support Grant",
     "Provides infrastructure support for research centers, similar to P30. Funds shared cores, pilot projects, and administrative infrastructure.",
     "5 years", "Varies"),

    ("P30",  "Center Grants",         "P-series",  "Center Core Support Grant",
     "Funds shared research cores (e.g., genomics, imaging, biostatistics, animal modeling) within an established research center. Does not directly fund research projects.",
     "5 years", "$500K–$2M+/yr"),

    ("P41",  "Center Grants",         "P-series",  "Biomedical Technology Research Resource",
     "Funds national centers that develop and disseminate cutting-edge biomedical technologies (e.g., cryo-EM, NMR, computing infrastructure) to the broader research community.",
     "5 years", "$1M–$5M/yr"),

    ("P50",  "Center Grants",         "P-series",  "Specialized Center Grant",
     "Large, multi-component research center focused on a specific disease or research area. Includes research projects, cores, and developmental projects. Similar to P01 but more center-focused.",
     "5 years", "$1M–$10M+/yr"),

    ("R00",  "Research Grants",       "R-series",  "Pathway to Independence Award — Phase II (Independent)",
     "Phase II of K99/R00. Awarded after the PI secures an independent faculty position. Provides research funding for the first years of independent research.",
     "3 years", "~$249K/yr direct costs"),

    ("R01",  "Research Grants",       "R-series",  "Research Project Grant",
     "The flagship NIH grant mechanism. Supports investigator-initiated, hypothesis-driven research. Most competitive and prestigious individual research grant. No budget cap (though $500K+ requires NCI council review).",
     "3–5 years", "$250K–$500K+/yr direct costs"),

    ("R03",  "Research Grants",       "R-series",  "Small Research Grant",
     "Provides limited funding for pilot studies, feasibility work, secondary data analysis, or small self-contained projects. No preliminary data required by some institutes.",
     "Up to 2 years", "Up to $50K/yr direct costs"),

    ("R13",  "Research Grants",       "R-series",  "Conference Support Grant",
     "Funds scientific conferences, symposia, and workshops that are relevant to NIH's mission. Cannot fund construction or equipment.",
     "1–5 years", "Up to $50K/yr"),

    ("R18",  "Research Grants",       "R-series",  "Research Demonstration and Dissemination Projects",
     "Supports projects that implement and evaluate evidence-based programs in real-world settings. Focuses on dissemination of research findings into practice.",
     "Up to 5 years", "Varies"),

    ("R21",  "Research Grants",       "R-series",  "Exploratory/Developmental Research Grant",
     "Funds high-risk, innovative ideas at an early stage. Typically used for proof-of-concept work before a full R01. No preliminary data often required.",
     "Up to 2 years", "Up to $275K total direct costs"),

    ("R24",  "Research Grants",       "R-series",  "Resource-Related Research Grant",
     "Funds the development of research resources — databases, repositories, reagents, model systems — that are shared with the broader scientific community.",
     "3–5 years", "Varies"),

    ("R25",  "Research Grants",       "R-series",  "Research Education Grant",
     "Supports research education programs: curriculum development, courses, workshops, short-term research experiences. Does not fund direct research.",
     "Up to 5 years", "Varies"),

    ("R33",  "Research Grants",       "R-series",  "Exploratory/Developmental Phase II Grant",
     "Phase II follow-on from R21. Supports continued development of a promising technology, tool, or assay that showed feasibility in the R21 phase.",
     "Up to 3 years", "Up to $500K/yr direct costs"),

    ("R34",  "Research Grants",       "R-series",  "Clinical Trial Planning Grant",
     "Supports planning and design activities needed before a full clinical trial. Covers protocol development, pilot testing, and regulatory preparation.",
     "Up to 3 years", "Up to $100K/yr direct costs"),

    ("R35",  "Research Grants",       "R-series",  "Outstanding Investigator Award",
     "Long-term funding for established, highly productive investigators. Provides stability and flexibility to pursue ambitious research without annual recompetition.",
     "7–8 years", "$600K–$1M+/yr direct costs"),

    ("R36",  "Research Grants",       "R-series",  "Dissertation Award",
     "NIH funding for pre-doctoral students to complete dissertation research. Provides stipend, research costs, and training-related expenses.",
     "Up to 2 years", "Up to $44K/yr"),

    ("R37",  "Research Grants",       "R-series",  "Method to Extend Research in Time (MERIT) Award",
     "Extended, no-competition renewal for R01 investigators with consistently outstanding reviews. Provides long-term research stability for top performers.",
     "Up to 10 years total", "Same as R01"),

    ("R38",  "Research Grants",       "R-series",  "Research Education Cooperative Agreement",
     "Supports research training and education programs, particularly those addressing workforce diversity in biomedical research.",
     "Up to 5 years", "Varies"),

    ("R49",  "Research Grants",       "R-series",  "Injury Prevention and Control Research Center Grant (CDC)",
     "CDC-funded mechanism supporting Injury Control Research Centers. Funds comprehensive injury prevention research, training, and outreach programs.",
     "5 years", "$500K–$1M+/yr"),

    ("R50",  "Research Grants",       "R-series",  "Research Specialist Award",
     "Provides long-term salary support for key research staff — bioinformaticians, statisticians, lab managers — embedded in research programs. Reduces dependence on soft money.",
     "5 years", "Salary + fringe benefits"),

    ("R56",  "Research Grants",       "R-series",  "High Priority, Short-Term Project Award (Bridge Award)",
     "Short-term emergency funding for high-scoring R01 applications that could not be funded due to budget constraints. Keeps the lab running during resubmission.",
     "1–2 years", "Same as R01 budget"),

    ("R61",  "Research Grants",       "R-series",  "Exploratory/Developmental Phase I (Milestone-Driven)",
     "Phase I of the R61/R33 pair. Milestone-driven: the team must achieve specific goals to advance to R33. Used for technology or method development.",
     "Up to 2 years", "Up to $275K/yr direct costs"),

    ("RF1",  "Research Grants",       "R-series",  "Multi-Year Funded Research Project (NIA/Alzheimer's)",
     "Used primarily by NIA for Alzheimer's Disease research. Functionally similar to R01 but awarded as a multi-year grant with full funding upfront.",
     "3–5 years", "Similar to R01"),

    ("S10",  "Shared Instrumentation","S-series",  "Shared Instrumentation Grant",
     "Funds purchase of expensive research equipment ($100K–$2M) that will be shared by multiple NIH-funded investigators at one institution. Covers MRI, mass spec, cryo-EM, etc.",
     "1-time award", "$100K–$2M"),

    ("T15",  "Training",              "T-series",  "Continuing Education Training Grant (Biomedical Informatics)",
     "Supports continuing education programs in health informatics and library science for health professionals.",
     "Up to 5 years", "Varies"),

    ("T32",  "Training",              "T-series",  "Institutional National Research Service Award",
     "Institutional training grant that funds multiple predoctoral and/or postdoctoral trainees. The institution selects trainees; NIH funds the program.",
     "5 years", "$50K–$200K+/yr depending on trainees"),

    ("T35",  "Training",              "T-series",  "Short-Term Institutional Research Training Grant",
     "Funds short-term (typically summer) research training experiences for medical, dental, or other health professional students.",
     "Up to 5 years", "Varies by # of trainees"),

    ("T42",  "Training",              "T-series",  "Occupational Safety and Health Training Program",
     "Funds training programs in occupational safety and health for professionals who will work in industrial or environmental health settings.",
     "5 years", "Varies"),

    ("U01",  "Cooperative Agreements","U-series",  "Research Project Cooperative Agreement",
     "Functionally similar to R01, but NIH program staff have substantial scientific involvement. Used for multi-site studies, large clinical trials, or when NIH wants programmatic oversight.",
     "3–5 years", "$250K–$1M+/yr"),

    ("U18",  "Cooperative Agreements","U-series",  "Research Demonstration Cooperative Agreement",
     "Cooperative version of R18. NIH has active involvement in research demonstration projects that test interventions in real-world settings.",
     "Up to 5 years", "Varies"),

    ("U19",  "Cooperative Agreements","U-series",  "Research Program Cooperative Agreement",
     "Large multi-project cooperative center where NIH has programmatic involvement. Similar to P01/P50 but with more NIH engagement. Often used for consortia.",
     "5 years", "$1M–$10M+/yr"),

    ("U24",  "Cooperative Agreements","U-series",  "Resource-Related Research Cooperative Agreement",
     "Cooperative version of R24. NIH co-manages the development of shared research resources — databases, reagent repositories, reference datasets.",
     "3–5 years", "Varies"),

    ("U2C",  "Cooperative Agreements","U-series",  "Trans-NIH Research Support Cooperative Agreement",
     "Cross-institute cooperative agreement supporting large-scale infrastructure or data initiatives that span multiple NIH institutes.",
     "5 years", "Varies"),

    ("U42",  "Cooperative Agreements","U-series",  "Develop and Distribute Animal Models & Related Materials",
     "Supports repositories that develop, maintain, and distribute animal models (e.g., transgenic mice, rats) or biological materials to the research community.",
     "5 years", "Varies"),

    ("U54",  "Cooperative Agreements","U-series",  "Specialized Center — Cooperative Agreement",
     "Cooperative version of P50. Large multi-component specialized research center with substantial NIH involvement. Often used for disease-focused consortia.",
     "5 years", "$2M–$15M+/yr"),

    ("UC2",  "Cooperative Agreements","U-series",  "Cooperative Agreement for Infrastructure",
     "Supports development of research infrastructure, platforms, or data systems under close NIH collaboration.",
     "Varies", "Varies"),

    ("UE5",  "Cooperative Agreements","U-series",  "Education Cooperative Agreement",
     "Supports research education and training programs with NIH programmatic involvement, often for building workforce diversity.",
     "Up to 5 years", "Varies"),

    ("UG1",  "Cooperative Agreements","U-series",  "Clinical Trial Cooperative Agreement",
     "Used for clinical trial networks where NIH coordinates multi-site trial infrastructure. Sites receive UG1 to participate in a network-wide protocol.",
     "5 years", "Varies per site"),

    ("UG3",  "Cooperative Agreements","U-series",  "Milestone-Driven Cooperative Agreement Phase I",
     "Phase I of UG3/UH3 pair. Milestone-driven planning/development phase for a large cooperative project. Must meet milestones to advance to UH3.",
     "Up to 2 years", "Varies"),

    ("UH3",  "Cooperative Agreements","U-series",  "Milestone-Driven Cooperative Agreement Phase II",
     "Phase II of UG3/UH3 pair. Full implementation phase after successful completion of UG3 milestones. Used for large clinical or research programs.",
     "Up to 4 years", "Varies"),

    ("UM1",  "Cooperative Agreements","U-series",  "Research Project with Complex Structure Cooperative Agreement",
     "Used for very large, complex clinical trials or research programs requiring coordinating centers. Often involves multi-institution consortia with centralized data management.",
     "5+ years", "$2M–$20M+/yr"),
]

# ── Load existing workbook ──────────────────────────────────────────────────
wb = load_workbook('jhu_grants_2024_2026.xlsx')

# Remove old tab if re-running
if 'Activity Codes' in wb.sheetnames:
    del wb['Activity Codes']

ws = wb.create_sheet('Activity Codes')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'A3'

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE  = 'FF1F3864'
MID_BLUE   = 'FF2E5090'
LIGHT_BLUE = 'FFD6E4F7'
WHITE      = 'FFFFFFFF'
GREY_ROW   = 'FFF5F7FA'

CAT_COLORS = {
    'Research Grants':        'FFDCE6F1',
    'Center Grants':          'FFE2EFDA',
    'Career Development':     'FFFFF2CC',
    'Fellowships':            'FFFCE4D6',
    'Training':               'FFE8D5F5',
    'Cooperative Agreements': 'FFD9EAD3',
    'Contracts':              'FFFFE599',
    'Director\'s Awards':     'FFCFE2F3',
    'Shared Instrumentation': 'FFD9D9D9',
    'Other Transactions':     'FFEAD1DC',
    'Other':                  'FFF2F2F2',
}

thin = Side(style='thin', color='FFB0BEC5')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Title banner ─────────────────────────────────────────────────────────
ws.merge_cells('A1:G1')
ws['A1'] = 'NIH Activity Codes — Reference Guide'
ws['A1'].font = Font(name='Arial', bold=True, size=13, color=WHITE)
ws['A1'].fill = PatternFill('solid', start_color=DARK_BLUE)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

# ── Header row ───────────────────────────────────────────────────────────
headers = ['Code', 'Category', 'Series', 'Full Name', 'Description', 'Typical Duration', 'Typical Award Size']
for col, hdr in enumerate(headers, 1):
    c = ws.cell(2, col, hdr)
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', start_color=MID_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[2].height = 22

# ── Load counts from data ────────────────────────────────────────────────
with open('jhu_grants_all.json') as f:
    raw = json.load(f)
counts = Counter(g.get('activity_code') for g in raw)

# ── Data rows ────────────────────────────────────────────────────────────
for row_i, (code, cat, series, name, desc, duration, size) in enumerate(
        sorted(CODES, key=lambda x: x[1]+x[0]), start=3):

    cat_color = CAT_COLORS.get(cat, 'FFF2F2F2')
    alt_fill  = PatternFill('solid', start_color=cat_color)
    row_fill  = alt_fill if row_i % 2 == 0 else PatternFill('solid', start_color=WHITE)

    vals = [code, cat, series, name, desc, duration, size]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row_i, col, val)
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True,
                                horizontal='center' if col in (1, 3, 6, 7) else 'left')
        if col == 1:
            c.font = Font(name='Arial', bold=True, size=10, color='FF1F3864')
            c.fill = PatternFill('solid', start_color=LIGHT_BLUE)
        elif col == 2:
            c.font = Font(name='Arial', bold=True, size=9)
            c.fill = alt_fill
        else:
            c.font = Font(name='Arial', size=9)
            c.fill = row_fill
    ws.row_dimensions[row_i].height = 42

# ── Column widths ─────────────────────────────────────────────────────────
widths = [8, 22, 12, 40, 80, 20, 24]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Save ──────────────────────────────────────────────────────────────────
wb.save('jhu_grants_2024_2026.xlsx')
print(f'Done — added Activity Codes tab with {len(CODES)} codes')
