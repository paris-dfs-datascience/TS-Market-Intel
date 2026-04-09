import re
import json
import warnings
warnings.filterwarnings('ignore')
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('jhu_nsf_grants_all.json') as f:
    raw = json.load(f)

ILLEGAL = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]')

def safe(val, default=''):
    if val is None:
        return default
    if isinstance(val, list):
        # flatten list of dicts or strings
        parts = []
        for v in val:
            if isinstance(v, dict):
                parts.append(v.get('name') or v.get('fullName') or json.dumps(v))
            else:
                parts.append(str(v))
        return ', '.join(parts)
    if isinstance(val, dict):
        return json.dumps(val)
    if isinstance(val, str):
        return ILLEGAL.sub('', val).strip()
    return val

def get_year(date_str):
    """Extract year from mm/dd/yyyy or yyyy-mm-dd."""
    if not date_str:
        return ''
    if '/' in str(date_str):
        parts = str(date_str).split('/')
        return parts[2] if len(parts) == 3 else ''
    return str(date_str)[:4]

def flatten(g):
    # co-PIs: can be list of dicts or string
    copi_raw = g.get('coPDPI') or []
    if isinstance(copi_raw, list):
        copis = ', '.join(
            c.get('fullName', '') if isinstance(c, dict) else str(c)
            for c in copi_raw
        )
    else:
        copis = str(copi_raw)

    return {
        'Award ID':               safe(g.get('id')),
        'Title':                  safe(g.get('title')),
        'Transaction Type':       safe(g.get('transType')),
        'Active':                 safe(g.get('activeAwd')),
        'Award Date':             safe(g.get('date')),
        'Start Date':             safe(g.get('startDate')),
        'End Date':               safe(g.get('expDate')),
        'Award Year':             get_year(g.get('date')),
        'Funds Obligated ($)':    int(g.get('fundsObligatedAmt') or 0),
        'Estimated Total ($)':    int(g.get('estimatedTotalAmt') or 0),
        'PI First Name':          safe(g.get('piFirstName')),
        'PI Last Name':           safe(g.get('piLastName')),
        'PI Email':               safe(g.get('piEmail')),
        'Co-PIs':                 safe(copis),
        'Program Director/PI':    safe(g.get('pdPIName')),
        'Program Officer':        safe(g.get('poName')),
        'Program Officer Email':  safe(g.get('poEmail')),
        'Organization':           safe(g.get('awardee')),
        'City':                   safe(g.get('awardeeCity')),
        'State':                  safe(g.get('awardeeStateCode')),
        'Zip Code':               safe(g.get('awardeeZipCode')),
        'UEI Number':             safe(g.get('ueiNumber')),
        'Perf City':              safe(g.get('perfCity')),
        'Perf State':             safe(g.get('perfStateCode')),
        'Directorate':            safe(g.get('dirAbbr')),
        'Division':               safe(g.get('divAbbr')),
        'Fund Program':           safe(g.get('fundProgramName')),
        'Program':                safe(g.get('program')),
        'Program Element Code':   safe(g.get('progEleCode')),
        'Program Ref Code':       safe(g.get('progRefCode')),
        'CFDA Number':            safe(g.get('cfdaNumber')),
        'Public Access Mandate':  safe(g.get('publicAccessMandate')),
        'Init Amendment Date':    safe(g.get('initAmendmentDate')),
        'Latest Amendment Date':  safe(g.get('latestAmendmentDate')),
        'Abstract':               safe(g.get('abstractText')),
        'Project Outcomes':       safe(g.get('projectOutComesReport')),
    }

records = [flatten(g) for g in raw]
headers = list(records[0].keys())

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_TEAL  = 'FF0D3349'
MID_TEAL   = 'FF1A5276'
LIGHT_TEAL = 'FFD1ECF1'
WHITE      = 'FFFFFFFF'
GREY_ROW   = 'FFF5F7FA'

hdr_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
hdr_fill  = PatternFill('solid', start_color=DARK_TEAL)
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=9)
thin      = Side(style='thin', color='FFB0BEC5')
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary Dashboard
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Summary Dashboard'
ws1.sheet_view.showGridLines = False
for col, w in zip('ABCDEF', [30, 18, 18, 18, 18, 18]):
    ws1.column_dimensions[col].width = w

# Title
ws1.merge_cells('A1:F1')
ws1['A1'] = 'Johns Hopkins University — NSF Grants Summary (2024–2026)'
ws1['A1'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
ws1['A1'].fill = PatternFill('solid', start_color=DARK_TEAL)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:F2')
ws1['A2'] = f'Source: NSF Awards API  |  Total Records: {len(records):,}  |  api.nsf.gov/services/v1/awards'
ws1['A2'].font = Font(name='Arial', italic=True, size=9, color='FF555555')
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 16

sub_font = Font(name='Arial', bold=True, color=WHITE, size=10)
sub_fill = PatternFill('solid', start_color=MID_TEAL)

def section_header(ws, row, title, span='A:F'):
    cols = span.split(':')
    ws.merge_cells(f'{cols[0]}{row}:{cols[1]}{row}')
    c = ws[f'{cols[0]}{row}']
    c.value = title
    c.font = sub_font
    c.fill = sub_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22

def col_headers(ws, row, hdrs):
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row, col, h)
        c.font = Font(name='Arial', bold=True, size=9, color='FF0D3349')
        c.fill = PatternFill('solid', start_color=LIGHT_TEAL)
        c.alignment = Alignment(horizontal='center')
        c.border = border
    ws.row_dimensions[row].height = 18

# ── By Year ──────────────────────────────────────────────────────────
r = 4
section_header(ws1, r, 'GRANTS BY AWARD YEAR')
r += 1
col_headers(ws1, r, ['Award Year', 'Grant Count', 'Total Obligated ($)', 'Avg Award ($)', 'Max Award ($)', 'Min Award ($)'])
data_start = r + 1

by_year = defaultdict(list)
for rec in records:
    by_year[rec['Award Year']].append(rec['Funds Obligated ($)'])

for i, yr in enumerate(sorted(by_year)):
    r += 1
    amts = [a for a in by_year[yr] if a > 0]
    fill = PatternFill('solid', start_color=WHITE if i % 2 == 0 else GREY_ROW)
    vals = [yr, len(by_year[yr]), sum(amts), round(sum(amts)/len(amts)) if amts else 0, max(amts) if amts else 0, min(amts) if amts else 0]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(r, col, v)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(horizontal='center')
        if col >= 3: c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 16

# Totals
r += 1
for col, v in enumerate(['TOTAL', f'=SUM(B{data_start}:B{r-1})', f'=SUM(C{data_start}:C{r-1})', '', '', ''], 1):
    c = ws1.cell(r, col, v)
    c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_TEAL)
    c.border = border; c.alignment = Alignment(horizontal='center')
    if col >= 3 and v: c.number_format = '$#,##0'
ws1.row_dimensions[r].height = 18

# ── By Transaction Type ───────────────────────────────────────────────
r += 2
section_header(ws1, r, 'BY TRANSACTION TYPE')
r += 1
col_headers(ws1, r, ['Transaction Type', 'Count', 'Total Obligated ($)', 'Avg Award ($)', '', ''])
by_type = defaultdict(list)
for rec in records:
    by_type[rec['Transaction Type']].append(rec['Funds Obligated ($)'])
for i, (tt, amts) in enumerate(sorted(by_type.items(), key=lambda x: -sum(x[1]))):
    r += 1
    clean = [a for a in amts if a > 0]
    fill = PatternFill('solid', start_color=WHITE if i % 2 == 0 else GREY_ROW)
    for col, v in enumerate([tt, len(amts), sum(clean), round(sum(clean)/len(clean)) if clean else 0, '', ''], 1):
        c = ws1.cell(r, col, v)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        if col >= 3 and v != '': c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 16

# ── By Directorate ────────────────────────────────────────────────────
r += 2
section_header(ws1, r, 'BY NSF DIRECTORATE')
r += 1
col_headers(ws1, r, ['Directorate', 'Division', 'Count', 'Total Obligated ($)', 'Avg Award ($)', ''])
by_dir = defaultdict(list)
for rec in records:
    key = (rec['Directorate'], rec['Division'])
    by_dir[key].append(rec['Funds Obligated ($)'])
for i, ((d, div), amts) in enumerate(sorted(by_dir.items(), key=lambda x: -sum(x[1]))):
    r += 1
    clean = [a for a in amts if a > 0]
    fill = PatternFill('solid', start_color=WHITE if i % 2 == 0 else GREY_ROW)
    for col, v in enumerate([d, div, len(amts), sum(clean), round(sum(clean)/len(clean)) if clean else 0, ''], 1):
        c = ws1.cell(r, col, v)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(horizontal='center')
        if col >= 4 and v != '': c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 16

# ── Top 5 by award ────────────────────────────────────────────────────
r += 2
section_header(ws1, r, 'TOP 5 GRANTS BY AWARD AMOUNT')
r += 1
col_headers(ws1, r, ['Award ID', 'PI', 'Directorate', 'Fund Program', 'Funds Obligated ($)', 'Title'])
top5 = sorted(records, key=lambda x: x['Funds Obligated ($)'], reverse=True)[:5]
for i, rec in enumerate(top5):
    r += 1
    fill = PatternFill('solid', start_color=WHITE if i % 2 == 0 else GREY_ROW)
    pi = f"{rec['PI First Name']} {rec['PI Last Name']}".strip()
    vals = [rec['Award ID'], pi, rec['Directorate'], rec['Fund Program'], rec['Funds Obligated ($)'], rec['Title'][:80]]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(r, col, v)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(horizontal='center' if col not in (2,4,6) else 'left', wrap_text=(col==6))
        if col == 5: c.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 30

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — All Grants
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('All Grants')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A2'

for col, hdr in enumerate(headers, 1):
    c = ws2.cell(1, col, hdr)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
ws2.row_dimensions[1].height = 30

for row_i, rec in enumerate(records, 2):
    fill = PatternFill('solid', start_color=WHITE if row_i % 2 == 0 else GREY_ROW)
    for col, key in enumerate(headers, 1):
        val = rec[key]
        c = ws2.cell(row_i, col, val)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=(key in ('Title', 'Abstract', 'Project Outcomes', 'Fund Program')))
        if key in ('Funds Obligated ($)', 'Estimated Total ($)'):
            c.number_format = '$#,##0'
    ws2.row_dimensions[row_i].height = 18

col_widths = {
    'Award ID': 12, 'Title': 55, 'Transaction Type': 22, 'Active': 8,
    'Award Date': 12, 'Start Date': 12, 'End Date': 12, 'Award Year': 10,
    'Funds Obligated ($)': 18, 'Estimated Total ($)': 18,
    'PI First Name': 14, 'PI Last Name': 14, 'PI Email': 26,
    'Co-PIs': 30, 'Program Director/PI': 22, 'Program Officer': 22,
    'Program Officer Email': 26, 'Organization': 30, 'City': 14, 'State': 8,
    'Zip Code': 10, 'UEI Number': 16, 'Perf City': 14, 'Perf State': 10,
    'Directorate': 12, 'Division': 12, 'Fund Program': 35, 'Program': 30,
    'Program Element Code': 16, 'Program Ref Code': 16, 'CFDA Number': 12,
    'Public Access Mandate': 10, 'Init Amendment Date': 16, 'Latest Amendment Date': 16,
    'Abstract': 70, 'Project Outcomes': 70,
}
for col, key in enumerate(headers, 1):
    ws2.column_dimensions[get_column_letter(col)].width = col_widths.get(key, 14)

# ════════════════════════════════════════════════════════════════════
# SHEET 3 — Top 25 by Award
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Top 25 by Award')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'A2'

t25_hdrs = ['Rank', 'Award ID', 'Award Year', 'Transaction Type', 'Funds Obligated ($)',
            'Estimated Total ($)', 'PI First Name', 'PI Last Name', 'Directorate',
            'Division', 'Fund Program', 'Title', 'Start Date', 'End Date']
for col, h in enumerate(t25_hdrs, 1):
    c = ws3.cell(1, col, h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
ws3.row_dimensions[1].height = 30

top25 = sorted(records, key=lambda x: x['Funds Obligated ($)'], reverse=True)[:25]
for row_i, rec in enumerate(top25, 2):
    fill = PatternFill('solid', start_color=WHITE if row_i % 2 == 0 else GREY_ROW)
    vals = [row_i - 1, rec['Award ID'], rec['Award Year'], rec['Transaction Type'],
            rec['Funds Obligated ($)'], rec['Estimated Total ($)'],
            rec['PI First Name'], rec['PI Last Name'],
            rec['Directorate'], rec['Division'], rec['Fund Program'],
            rec['Title'], rec['Start Date'], rec['End Date']]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row_i, col, v)
        c.font = cell_font; c.fill = fill; c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=(t25_hdrs[col-1] in ('Title', 'Fund Program')),
                                horizontal='center' if col not in (4, 11, 12) else 'left')
        if t25_hdrs[col-1] in ('Funds Obligated ($)', 'Estimated Total ($)'):
            c.number_format = '$#,##0'
    ws3.row_dimensions[row_i].height = 35

t25_widths = [6, 12, 10, 22, 18, 18, 14, 14, 12, 12, 35, 55, 12, 12]
for col, w in enumerate(t25_widths, 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════
out = 'jhu_nsf_grants_2024_2026.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Sheets: Summary Dashboard | All Grants ({len(records)} rows) | Top 25 by Award')
