import json
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT  = "/Users/amanpreet/Desktop/thomas-scientific-market-intelligence/80s-accounts-market-intelligence/backend/accounts-vertical/unified_signals.json"
OUTPUT = "/Users/amanpreet/Desktop/thomas-scientific-market-intelligence/80s-accounts-market-intelligence/backend/accounts-vertical/unified_signals.xlsx"

# ── colours ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
WHITE      = "FFFFFF"
GRAY_BG    = "D9D9D9"
LIGHT_BLUE = "DDEEFF"
LIGHT_ORG  = "FDEBD0"
LIGHT_GRN  = "DDFFDD"
YELLOW_HL  = "FFFACD"
ALT_ROW    = "F5F8FF"

SRC_COLORS = {"80s_accounts": LIGHT_BLUE, "NIH": LIGHT_ORG, "NSF": LIGHT_GRN}

STYPE_COLORS = {
    "grant": "E8F4FD", "contract": "FEF9E7", "expansion": "E8F8F5",
    "project": "F4ECF7", "tender": "FDFEFE", "capital": "FEF5E7",
    "pipeline": "EBF5FB", "funding": "E9F7EF", "regulatory": "FDEDEC",
    "hiring": "F9EBEA", "partnership": "EAF2FF", "faculty": "FFF3E0",
}

CAT_COLORS = {
    "Government":              "C8DCF5",
    "BioPharma":               "F5C8DC",
    "Education & Research":    "C8F5D8",
    "Hospital & Health Systems": "F5EBC8",
}

# ── helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def parse_amount(s):
    if not s:
        return 0.0
    cleaned = re.sub(r"[^\d.]", "", str(s).replace(",", ""))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header(ws, row, headers, bg=DARK_BLUE, fg=WHITE, height=20):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = font(bold=True, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border()
    ws.row_dimensions[row].height = height

def body_cell(ws, row, col, value, bg=WHITE, bold=False, halign="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=halign, vertical="top", wrap_text=wrap)
    c.border = border()
    return c

# ── load data ─────────────────────────────────────────────────────────────────
with open(INPUT) as f:
    data = json.load(f)

generated_at   = data["generated_at"]
total_signals  = data["total_signals"]
sources        = data["sources"]
accounts_dict  = data["accounts"]
all_signals    = data["all_signals"]

acc_src_counts = {}
for name, acc in accounts_dict.items():
    cnt = defaultdict(int)
    for sig in acc["signals"]:
        cnt[sig["source"]] += 1
    acc_src_counts[name] = cnt

cat_stats = defaultdict(lambda: {"accounts": set(), "signals": 0, "sources": set()})
for name, acc in accounts_dict.items():
    cat = acc["category"]
    cat_stats[cat]["accounts"].add(name)
    cat_stats[cat]["signals"] += acc["signal_count"]
    for src in acc["sources"]:
        cat_stats[cat]["sources"].add(src)

top20     = sorted(accounts_dict.values(), key=lambda a: a["signal_count"], reverse=True)[:20]
multi_src = {n: a for n, a in accounts_dict.items() if len(a["sources"]) >= 2}

wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard")

def big_merge(ws, row, end_col, text, bg, fnt, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = height

big_merge(ws, 1, 9,
          "Thomas Scientific  //  Market Intelligence — Unified View",
          DARK_BLUE,
          font(bold=True, size=14, color=WHITE), height=36)

big_merge(ws, 2, 9,
          f"Sources: 80s Accounts (Gemini) + NIH Grants + NSF Grants     |     Generated: {generated_at[:10]}",
          "4472C4",
          font(italic=True, size=10, color=WHITE), height=20)

ws.row_dimensions[3].height = 8

# Source summary box rows 4-8
box = [
    ("Source",       "Signal Count", DARK_BLUE, True,  WHITE),
    ("80s Accounts", sources["80s_accounts"], LIGHT_BLUE, False, "000000"),
    ("NIH Grants",   sources["NIH"],          LIGHT_ORG,  False, "000000"),
    ("NSF Grants",   sources["NSF"],          LIGHT_GRN,  False, "000000"),
    ("Total",        total_signals,           GRAY_BG,    True,  "000000"),
]
for i, (label, count, bg_color, bold, fg) in enumerate(box):
    r = 4 + i
    for j, v in enumerate([label, count], 2):
        c = ws.cell(row=r, column=j, value=v)
        c.font = font(bold=bold, color=fg)
        c.fill = fill(bg_color)
        c.alignment = Alignment(horizontal="left" if j == 2 else "right", vertical="center")
        c.border = border()
    ws.row_dimensions[r].height = 18
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 14

ws.row_dimensions[9].height = 8

# Category summary — row 10
r = 10
big_merge(ws, r, 6, "Category Summary", DARK_BLUE,
          font(bold=True, size=11, color=WHITE), height=22)
r += 1
write_header(ws, r, ["Category", "Accounts", "Signals", "Sources"], bg="2F5496")
r += 1

for cat in sorted(cat_stats.keys()):
    stats = cat_stats[cat]
    bg_c = CAT_COLORS.get(cat, "F0F0F0")
    vals = [cat, len(stats["accounts"]), stats["signals"], ", ".join(sorted(stats["sources"]))]
    for j, v in enumerate(vals, 1):
        body_cell(ws, r, j, v, bg=bg_c, halign="left" if j in (1,4) else "center")
    ws.row_dimensions[r].height = 18
    r += 1

ws.row_dimensions[r].height = 8
r += 1

# Top-20 table
big_merge(ws, r, 9, "Top 20 Accounts by Signal Count", DARK_BLUE,
          font(bold=True, size=11, color=WHITE), height=22)
r += 1
t20_hdrs = ["Account", "Category", "Total Signals", "80s Signals",
            "NIH Signals", "NSF Signals", "Multi-Source?", "Sources"]
write_header(ws, r, t20_hdrs, bg="2F5496", height=28)
r += 1

for ri, acc in enumerate(top20):
    sc    = acc_src_counts[acc["account"]]
    multi = len(acc["sources"]) >= 2
    bg_r  = ALT_ROW if ri % 2 == 1 else WHITE
    bold  = multi

    vals = [acc["account"], acc["category"], acc["signal_count"],
            sc.get("80s_accounts", 0), sc.get("NIH", 0), sc.get("NSF", 0),
            "Yes" if multi else "No",
            ", ".join(sorted(acc["sources"]))]

    for j, v in enumerate(vals, 1):
        bg_use = (LIGHT_BLUE if j == 4 else
                  LIGHT_ORG  if j == 5 else
                  LIGHT_GRN  if j == 6 else
                  YELLOW_HL  if (j == 7 and multi) else
                  bg_r)
        body_cell(ws, r, j, v, bg=bg_use, bold=bold,
                  halign="left" if j in (1,2,8) else "center")
    ws.row_dimensions[r].height = 18
    r += 1

# dashboard column widths (A through I)
col_width(ws, [2, 40, 24, 14, 12, 12, 12, 14, 30])

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — All Signals (≤5000)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Signals")

sorted_sigs = sorted(
    all_signals,
    key=lambda s: (s.get("category",""), s.get("account",""), s.get("source",""))
)[:5000]

hdrs2 = ["Account", "Category", "Source", "Signal Type", "Summary",
         "Why It Matters", "Amount", "PI", "Start Date", "End Date"]
write_header(ws2, 1, hdrs2)

for ri, sig in enumerate(sorted_sigs):
    r    = ri + 2
    src  = sig.get("source","")
    stype = sig.get("signal_type","")
    row_bg = SRC_COLORS.get(src, WHITE)

    row_vals = [sig.get("account",""), sig.get("category",""), src, stype,
                sig.get("summary",""), sig.get("why_it_matters",""),
                sig.get("amount",""), sig.get("pi",""),
                sig.get("start_date",""), sig.get("end_date","")]

    for j, v in enumerate(row_vals, 1):
        bg_use = (SRC_COLORS.get(src, WHITE)   if j == 3 else
                  STYPE_COLORS.get(stype, WHITE) if j == 4 else
                  row_bg)
        wrap = j in (5, 6)
        body_cell(ws2, r, j, v, bg=bg_use, wrap=wrap)
    ws2.row_dimensions[r].height = 50

col_width(ws2, [32, 22, 14, 14, 55, 55, 14, 22, 12, 12])
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — By Category
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("By Category")

r = 1
big_merge(ws3, r, 6, "Accounts by Category — Signal Intelligence Breakdown",
          DARK_BLUE, font(bold=True, size=13, color=WHITE), height=28)
r += 1

for cat in sorted(cat_stats.keys()):
    ws3.row_dimensions[r].height = 8
    r += 1

    # category header bar
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws3.cell(row=r, column=1, value=cat)
    c.font = font(bold=True, size=11, color=WHITE)
    c.fill = fill(CAT_COLORS.get(cat, "4472C4"))
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[r].height = 22
    r += 1

    write_header(ws3, r, ["Account", "Total Signals", "80s Signals",
                           "NIH Signals", "NSF Signals", "Sources"], bg="2F5496")
    r += 1

    cat_accs = sorted([a for a in accounts_dict.values() if a["category"] == cat],
                      key=lambda a: a["signal_count"], reverse=True)
    for ri2, acc in enumerate(cat_accs):
        sc = acc_src_counts[acc["account"]]
        bg_r = ALT_ROW if ri2 % 2 == 1 else WHITE
        vals = [acc["account"], acc["signal_count"],
                sc.get("80s_accounts",0), sc.get("NIH",0), sc.get("NSF",0),
                ", ".join(sorted(acc["sources"]))]
        for j, v in enumerate(vals, 1):
            bg_use = (LIGHT_BLUE if j == 3 else
                      LIGHT_ORG  if j == 4 else
                      LIGHT_GRN  if j == 5 else
                      bg_r)
            body_cell(ws3, r, j, v, bg=bg_use,
                      halign="left" if j in (1,6) else "center")
        ws3.row_dimensions[r].height = 18
        r += 1

col_width(ws3, [40, 14, 12, 12, 12, 30])
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Multi-Source Accounts
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Multi-Source Accounts")

big_merge(ws4, 1, 7,
          "Highest Confidence Accounts — Multiple Intelligence Sources",
          DARK_BLUE, font(bold=True, size=13, color=WHITE), height=28)

write_header(ws4, 2, ["Account", "Category", "Total", "80s Signals",
                       "NIH Signals", "NSF Signals", "Signal Types"])

ms_sorted = sorted(multi_src.values(), key=lambda a: a["signal_count"], reverse=True)
for ri, acc in enumerate(ms_sorted):
    r  = ri + 3
    sc = acc_src_counts[acc["account"]]
    stypes = sorted(set(s["signal_type"] for s in acc["signals"]))
    vals   = [acc["account"], acc["category"], acc["signal_count"],
              sc.get("80s_accounts",0), sc.get("NIH",0), sc.get("NSF",0),
              ", ".join(stypes)]
    for j, v in enumerate(vals, 1):
        bg_use = (LIGHT_BLUE if j == 4 else
                  LIGHT_ORG  if j == 5 else
                  LIGHT_GRN  if j == 6 else
                  YELLOW_HL)
        body_cell(ws4, r, j, v, bg=bg_use,
                  halign="left" if j in (1,2,7) else "center")
    ws4.row_dimensions[r].height = 18

col_width(ws4, [38, 24, 10, 12, 12, 12, 40])
ws4.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — NIH Top Grants
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("NIH Top Grants")

nih_top = sorted(
    [s for s in all_signals if s["source"] == "NIH"],
    key=lambda s: parse_amount(s.get("amount","")), reverse=True
)[:100]

big_merge(ws5, 1, 9, "NIH Top 100 Grants by Award Amount",
          "8B0000", font(bold=True, size=13, color=WHITE), height=28)
write_header(ws5, 2, ["Account", "Category", "Project Title", "PI", "Amount ($)",
                       "Signal Type", "Start Date", "End Date", "Source URL"],
             bg="C0392B")

for ri, sig in enumerate(nih_top):
    r      = ri + 3
    bg_r   = ALT_ROW if ri % 2 == 1 else WHITE
    amt_v  = parse_amount(sig.get("amount",""))
    vals   = [sig.get("account",""), sig.get("category",""),
              sig.get("title","") or sig.get("summary",""),
              sig.get("pi",""),
              amt_v if amt_v else sig.get("amount",""),
              sig.get("signal_type",""),
              sig.get("start_date",""), sig.get("end_date",""),
              sig.get("source_url","")]
    for j, v in enumerate(vals, 1):
        c = body_cell(ws5, r, j, v, bg=bg_r, wrap=(j==3))
        if j == 5 and isinstance(v, float) and v > 0:
            c.number_format = '$#,##0'
    ws5.row_dimensions[r].height = 40

col_width(ws5, [32, 22, 55, 25, 14, 12, 12, 12, 50])
ws5.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — NSF Top Grants
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("NSF Top Grants")

nsf_top = sorted(
    [s for s in all_signals if s["source"] == "NSF"],
    key=lambda s: parse_amount(s.get("amount","")), reverse=True
)[:100]

big_merge(ws6, 1, 9, "NSF Top 100 Grants by Award Amount",
          "155724", font(bold=True, size=13, color=WHITE), height=28)
write_header(ws6, 2, ["Account", "Category", "Project Title", "PI", "Amount ($)",
                       "Signal Type", "Start Date", "End Date", "Source URL"],
             bg="1E7E34")

for ri, sig in enumerate(nsf_top):
    r     = ri + 3
    bg_r  = ALT_ROW if ri % 2 == 1 else WHITE
    amt_v = parse_amount(sig.get("amount",""))
    vals  = [sig.get("account",""), sig.get("category",""),
             sig.get("title","") or sig.get("summary",""),
             sig.get("pi",""),
             amt_v if amt_v else sig.get("amount",""),
             sig.get("signal_type",""),
             sig.get("start_date",""), sig.get("end_date",""),
             sig.get("source_url","")]
    for j, v in enumerate(vals, 1):
        c = body_cell(ws6, r, j, v, bg=bg_r, wrap=(j==3))
        if j == 5 and isinstance(v, float) and v > 0:
            c.number_format = '$#,##0'
    ws6.row_dimensions[r].height = 40

col_width(ws6, [32, 22, 55, 25, 14, 12, 12, 12, 50])
ws6.freeze_panes = "A3"

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")

import os
mb = os.path.getsize(OUTPUT) / 1024 / 1024
print(f"File size: {mb:.2f} MB  ({os.path.getsize(OUTPUT):,} bytes)")
