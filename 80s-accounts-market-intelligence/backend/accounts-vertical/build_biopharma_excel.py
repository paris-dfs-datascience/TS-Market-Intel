import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── constants ────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EBF3FB"
LIGHT_GRAY = "F2F2F2"
GRAY_TEXT  = "666666"

SIGNAL_TYPES = ["grant","capital","contract","pipeline","expansion",
                "partnership","funding","project","regulatory","hiring"]

SIGNAL_COLORS = {
    "grant":       "DDEEFF",
    "capital":     "DDFFDD",
    "contract":    "FFFACC",
    "pipeline":    "E8DAEF",
    "expansion":   "D5F5E3",
    "partnership": "FADBD8",
    "funding":     "FDEBD0",
    "project":     "FFE5CC",
    "regulatory":  "FDEDEC",
    "hiring":      "EBF5FB",
    "tender":      "F0F0F0",
}

# best metadata fields per signal type → (Detail1 label, Detail2 label)
DETAIL_FIELDS = {
    "grant":       ("Agency / Recipient",         "Amount"),
    "capital":     ("Project Name",               "Value / Location"),
    "contract":    ("Contract Name",              "Est. Value"),
    "pipeline":    ("Product / Program",          "Stage"),
    "expansion":   ("Location",                   "Investment Value"),
    "partnership": ("Partner",                    "Deal Type / Value"),
    "funding":     ("Amount",                     "Funding Type"),
    "project":     ("Project Name",               "Timeline"),
    "regulatory":  ("Product / Site",             "Regulatory Action"),
    "hiring":      ("Role / Department",          "Location / Headcount"),
    "tender":      ("Detail 1",                   "Detail 2"),
}

def get_detail_values(sig_type, signal):
    if sig_type == "grant":
        d1 = f"{signal.get('agency','')} / {signal.get('recipient','')}".strip(" /")
        d2 = signal.get('amount', '')
    elif sig_type == "capital":
        d1 = signal.get('project_name', '')
        loc = signal.get('location', '')
        val = signal.get('value', '')
        d2 = f"{val} | {loc}".strip(" |") if (val or loc) else ''
    elif sig_type == "contract":
        d1 = signal.get('contract_name', '')
        d2 = signal.get('estimated_value', '')
    elif sig_type == "pipeline":
        d1 = signal.get('product_or_program', '')
        d2 = signal.get('stage', '')
    elif sig_type == "expansion":
        d1 = signal.get('location', '')
        d2 = signal.get('investment_value', '')
    elif sig_type == "partnership":
        d1 = signal.get('partner', '')
        dt = signal.get('deal_type', '')
        dv = signal.get('deal_value', '')
        d2 = f"{dt} | {dv}".strip(" |") if (dt or dv) else ''
    elif sig_type == "funding":
        d1 = signal.get('amount', '')
        d2 = signal.get('funding_type', '')
    elif sig_type == "project":
        d1 = signal.get('project_name', '')
        d2 = signal.get('timeline', '')
    elif sig_type == "regulatory":
        d1 = signal.get('product_or_site', '')
        d2 = signal.get('regulatory_action', '')
    elif sig_type == "hiring":
        role = signal.get('role_or_department', '')
        loc  = signal.get('location', '')
        hc   = signal.get('headcount', '')
        d1   = role
        d2   = f"{loc} | HC: {hc}".strip(" |") if (loc or hc) else ''
    else:
        d1 = d2 = ''
    return str(d1) if d1 else '', str(d2) if d2 else ''

def hdr_font(size=10, bold=True):
    return Font(name="Arial", bold=bold, color=WHITE, size=size)

def hdr_fill():
    return PatternFill("solid", fgColor=DARK_BLUE)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = border
    if number_format: c.number_format = number_format
    return c

# ── load data ─────────────────────────────────────────────────────────────────
with open('/Users/amanpreet/Desktop/thomas-scientific-market-intelligence/80s-accounts-market-intelligence/backend/accounts-vertical/biopharma_top12_results.json') as f:
    data = json.load(f)

wb = Workbook()
wb.remove(wb.active)

# ── pre-compute counts ────────────────────────────────────────────────────────
account_rows = []   # (account, category, {sig_type: count}, total)
for entry in data:
    acct  = entry['account']
    cat   = entry['category']
    sigs  = entry['signals']
    cnts  = {st: len(sigs.get(st, [])) for st in SIGNAL_TYPES}
    total = sum(cnts.values())
    account_rows.append((acct, cat, cnts, total))

account_rows.sort(key=lambda x: -x[3])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Summary")

# Title (row 1) merged across all 12 cols
ws1.merge_cells("A1:L1")
c = ws1["A1"]
c.value     = "Thomas Scientific // BioPharma Market Intelligence — Top 12 Accounts"
c.font      = Font(name="Arial", bold=True, color=WHITE, size=12)
c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

# Subtitle (row 2)
ws1.merge_cells("A2:L2")
c = ws1["A2"]
c.value     = "Last 7 Days | April 7, 2026"
c.font      = Font(name="Arial", italic=True, color=GRAY_TEXT, size=10)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

# Blank row 3
ws1.row_dimensions[3].height = 8

# Header row 4
headers = ["Account","Total Signals","Grant","Capital","Contract","Pipeline",
           "Expansion","Partnership","Funding","Project","Regulatory","Hiring"]
for col, h in enumerate(headers, 1):
    write_cell(ws1, 4, col, h,
               font=Font(name="Arial", bold=True, color=WHITE, size=10),
               fill=PatternFill("solid", fgColor=DARK_BLUE),
               alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=thin_border())
ws1.row_dimensions[4].height = 24

# Data rows (start row 5)
DATA_START = 5
col_letters = {st: get_column_letter(i+3) for i, st in enumerate(SIGNAL_TYPES)}  # C..L

for i, (acct, cat, cnts, total) in enumerate(account_rows):
    row  = DATA_START + i
    fill_color = LIGHT_BLUE if i % 2 == 1 else WHITE
    row_fill   = PatternFill("solid", fgColor=fill_color)
    bold_flag  = i < 3

    # Account
    write_cell(ws1, row, 1, acct,
               font=Font(name="Arial", bold=bold_flag, size=10),
               fill=row_fill,
               alignment=Alignment(horizontal="left", vertical="center"),
               border=thin_border())

    # Total Signals — SUM formula across C..L
    sum_formula = f"=SUM(C{row}:L{row})"
    write_cell(ws1, row, 2, sum_formula,
               font=Font(name="Arial", bold=bold_flag, size=10),
               fill=row_fill,
               alignment=Alignment(horizontal="center", vertical="center"),
               border=thin_border())

    # Each signal type count
    for j, st in enumerate(SIGNAL_TYPES):
        col = j + 3
        write_cell(ws1, row, col, cnts[st],
                   font=Font(name="Arial", bold=bold_flag, size=10),
                   fill=row_fill,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=thin_border())

ws1.row_dimensions[row].height = 18  # set last row height as default

# Totals row
n_data = len(account_rows)
tot_row = DATA_START + n_data
tot_fill = PatternFill("solid", fgColor="D9D9D9")
write_cell(ws1, tot_row, 1, "TOTALS",
           font=Font(name="Arial", bold=True, size=10),
           fill=tot_fill,
           alignment=Alignment(horizontal="left", vertical="center"),
           border=thin_border())
# Total Signals column
write_cell(ws1, tot_row, 2,
           f"=SUM(C{tot_row}:L{tot_row})",
           font=Font(name="Arial", bold=True, size=10),
           fill=tot_fill,
           alignment=Alignment(horizontal="center", vertical="center"),
           border=thin_border())
for j in range(len(SIGNAL_TYPES)):
    col     = j + 3
    col_ltr = get_column_letter(col)
    write_cell(ws1, tot_row, col,
               f"=SUM({col_ltr}{DATA_START}:{col_ltr}{tot_row-1})",
               font=Font(name="Arial", bold=True, size=10),
               fill=tot_fill,
               alignment=Alignment(horizontal="center", vertical="center"),
               border=thin_border())
ws1.row_dimensions[tot_row].height = 20

# Column widths
ws1.column_dimensions["A"].width = 35
for j in range(len(SIGNAL_TYPES) + 1):   # B..L (Total Signals + 10 types)
    ws1.column_dimensions[get_column_letter(j + 2)].width = 10

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — All Signals
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Signals")

all_headers = ["Account","Category","Signal Type","Summary","Why It Matters","Source URL","Detail 1","Detail 2"]
for col, h in enumerate(all_headers, 1):
    write_cell(ws2, 1, col, h,
               font=Font(name="Arial", bold=True, color=WHITE, size=10),
               fill=PatternFill("solid", fgColor=DARK_BLUE),
               alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=thin_border())
ws2.row_dimensions[1].height = 22
ws2.freeze_panes = "A2"

row = 2
for entry in data:
    acct = entry['account']
    cat  = entry['category']
    for st in SIGNAL_TYPES:
        for signal in entry['signals'].get(st, []):
            sig_fill = PatternFill("solid", fgColor=SIGNAL_COLORS.get(st, "FFFFFF"))
            d1, d2 = get_detail_values(st, signal)
            row_data = [
                acct,
                cat,
                st,
                signal.get('summary', ''),
                signal.get('why_it_matters', ''),
                signal.get('source_url', ''),
                d1,
                d2,
            ]
            for col, val in enumerate(row_data, 1):
                wrap = col in (4, 5)
                aln  = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
                write_cell(ws2, row, col, val,
                           font=Font(name="Arial", size=9),
                           fill=sig_fill if col == 3 else PatternFill("solid", fgColor="FFFFFF"),
                           alignment=aln,
                           border=thin_border())
            ws2.row_dimensions[row].height = 60
            row += 1

# Column widths
for col, width in zip(range(1, 9), [25, 15, 12, 60, 50, 40, 20, 20]):
    ws2.column_dimensions[get_column_letter(col)].width = width

# ═════════════════════════════════════════════════════════════════════════════
# SHEETS 3+ — Per-Account sheets
# ═════════════════════════════════════════════════════════════════════════════
for entry in data:
    acct     = entry['account']
    cat      = entry['category']
    sigs     = entry['signals']
    total    = sum(len(v) for v in sigs.values())
    if total == 0:
        continue

    sheet_name = acct[:31]
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = acct
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = f"{cat}   |   Total Signals: {total}"
    c.font      = Font(name="Arial", italic=True, color=GRAY_TEXT, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    current_row = 3

    for st in SIGNAL_TYPES:
        signals = sigs.get(st, [])
        if not signals:
            continue

        # Blank spacer
        current_row += 1

        # Section header
        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws[f"A{current_row}"]
        c.value     = f"  {st.upper()}  ({len(signals)} signal{'s' if len(signals)>1 else ''})"
        c.font      = Font(name="Arial", bold=True, color="1F3864", size=10)
        c.fill      = PatternFill("solid", fgColor=SIGNAL_COLORS.get(st, "F0F0F0"))
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Column headers
        d1_lbl, d2_lbl = DETAIL_FIELDS.get(st, ("Detail 1", "Detail 2"))
        col_hdrs = ["#", "Summary", "Why It Matters", "Source URL", d1_lbl, d2_lbl, ""]
        for col, h in enumerate(col_hdrs[:6], 1):
            write_cell(ws, current_row, col, h,
                       font=Font(name="Arial", bold=True, color=WHITE, size=9),
                       fill=PatternFill("solid", fgColor=DARK_BLUE),
                       alignment=Alignment(horizontal="center", vertical="center"),
                       border=thin_border())
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Signal rows
        for idx, signal in enumerate(signals, 1):
            d1, d2 = get_detail_values(st, signal)
            row_fill = PatternFill("solid", fgColor=LIGHT_BLUE if idx % 2 == 0 else WHITE)
            vals = [idx,
                    signal.get('summary', ''),
                    signal.get('why_it_matters', ''),
                    signal.get('source_url', ''),
                    d1, d2, '']
            for col, val in enumerate(vals[:6], 1):
                wrap = col in (2, 3)
                write_cell(ws, current_row, col, val,
                           font=Font(name="Arial", size=9),
                           fill=row_fill,
                           alignment=Alignment(horizontal="left" if col > 1 else "center",
                                               vertical="top", wrap_text=wrap),
                           border=thin_border())
            ws.row_dimensions[current_row].height = 60
            current_row += 1

    # Column widths
    widths = [4, 60, 50, 40, 22, 22, 5]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ── save ─────────────────────────────────────────────────────────────────────
output_path = ('/Users/amanpreet/Desktop/thomas-scientific-market-intelligence/'
               '80s-accounts-market-intelligence/backend/accounts-vertical/'
               'biopharma_top12_signals.xlsx')
wb.save(output_path)
print(f"Saved: {output_path}")

import os
size_kb = os.path.getsize(output_path) / 1024
print(f"File size: {size_kb:.1f} KB")
